#!/usr/bin/env python3
"""
Build publication-ready Supplementary Table S1 for the ancient-DNA
cross-library analysis.

Inputs
------
1. AncientMetagenomeDir environmental-library metadata TSV.
2. best_length_per_library.csv from the period-3 aggregation.

Matching
--------
- First match analysis sample -> AncientMetagenomeDir archive_data_accession.
- If no direct match exists, match the analysis accession against accessions
  embedded in AncientMetagenomeDir download_links.
- Both identifiers are retained when they differ.

Outputs
-------
- Supplementary_Table_S1.xlsx
    * README sheet
    * Supplementary Table S1 sheet
- Supplementary_Table_S1.tsv
- Supplementary_Table_S1_match_report.tsv

Requires: pandas, openpyxl
"""

from __future__ import annotations

import argparse
import hashlib
import re
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ACCESSION_RE = re.compile(r"\b(?:ERR|SRR|DRR)\d+\b", re.IGNORECASE)

OUTPUT_COLUMNS = [
    "analysis_accession",
    "AncientMetagenomeDir_accession",
    "match_method",
    "project_name",
    "publication_year",
    "publication_doi",
    "sample_name",
    "archive",
    "archive_project",
    "archive_sample_accession",
    "library_name",
    "strand_type",
    "library_polymerase",
    "library_treatment",
    "library_concentration",
    "instrument_model",
    "library_layout",
    "library_strategy",
    "read_used",
    "reported_read_count",
    "best_length_bp",
    "max_mean_period3_R2",
    "download_links",
    "download_md5s",
    "download_sizes",
    "notes",
]

COLUMN_DESCRIPTIONS = {
    "analysis_accession":
        "Accession/sample identifier used by the period-3 analysis; for downloaded FASTQ data this is derived from the FASTQ filename.",
    "AncientMetagenomeDir_accession":
        "Value recorded in the AncientMetagenomeDir archive_data_accession field.",
    "match_method":
        "How the analysis row was linked to AncientMetagenomeDir: direct_accession or download_link_accession.",
    "project_name": "AncientMetagenomeDir project/study label.",
    "publication_year": "Publication year reported by AncientMetagenomeDir.",
    "publication_doi": "DOI of the source data publication.",
    "sample_name": "Sample name reported by AncientMetagenomeDir.",
    "archive": "Sequence archive reported by AncientMetagenomeDir.",
    "archive_project": "Archive project accession.",
    "archive_sample_accession": "Archive sample accession.",
    "library_name": "Library identifier reported by AncientMetagenomeDir.",
    "strand_type": "Library strand type reported by AncientMetagenomeDir.",
    "library_polymerase": "Library polymerase reported by AncientMetagenomeDir.",
    "library_treatment": "Library treatment reported by AncientMetagenomeDir.",
    "library_concentration": "Library concentration field reported by AncientMetagenomeDir.",
    "instrument_model": "Sequencing instrument model.",
    "library_layout": "SINGLE or PAIRED library layout.",
    "library_strategy": "Sequencing strategy.",
    "read_used":
        "Read used for the read-coordinate analysis: R1 for paired-end datasets; single read for single-end datasets.",
    "reported_read_count": "Read count reported in the AncientMetagenomeDir metadata.",
    "best_length_bp":
        "Exact read length selected for the cross-library figure.",
    "max_mean_period3_R2":
        "Largest mean period-3 R² across eligible exact read lengths; within each length the mean is across A, T, G, and C.",
    "download_links": "Raw-data download link(s) recorded by AncientMetagenomeDir.",
    "download_md5s": "MD5 checksum(s) corresponding to the download link(s).",
    "download_sizes": "Download size(s) corresponding to the download link(s).",
    "notes": "Notes on accession discrepancies or other matching details.",
}


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def accessions_from_download_links(value) -> set[str]:
    if pd.isna(value):
        return set()
    return {m.upper() for m in ACCESSION_RE.findall(str(value))}


def read_used(layout) -> str:
    text = "" if pd.isna(layout) else str(layout).strip().upper()
    if text == "PAIRED":
        return "R1"
    if text == "SINGLE":
        return "single read"
    return "not specified"


def require_columns(df: pd.DataFrame, required: Iterable[str], name: str) -> None:
    missing = sorted(set(required) - set(df.columns))
    if missing:
        raise ValueError(f"{name} is missing required columns: {missing}")


def build_table(meta: pd.DataFrame, results: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    require_columns(
        meta,
        [
            "project_name", "publication_year", "data_publication_doi",
            "sample_name", "archive", "archive_project",
            "archive_sample_accession", "library_name", "strand_type",
            "library_polymerase", "library_treatment",
            "library_concentration", "instrument_model", "library_layout",
            "library_strategy", "read_count", "archive_data_accession",
            "download_links", "download_md5s", "download_sizes",
        ],
        "AncientMetagenomeDir metadata",
    )
    require_columns(
        results, ["sample", "category", "best_length", "max_r2"],
        "best_length_per_library",
    )

    meta = meta.copy()
    results = results.copy()
    meta["archive_data_accession_norm"] = (
        meta["archive_data_accession"].astype("string").str.strip().str.upper()
    )
    results["analysis_accession_norm"] = (
        results["sample"].astype("string").str.strip().str.upper()
    )

    # Guard against ambiguous direct accession matches.
    dup = meta["archive_data_accession_norm"].dropna()
    dup = dup[dup.duplicated(keep=False)]
    if not dup.empty:
        raise ValueError(
            "AncientMetagenomeDir archive_data_accession is not unique for: "
            + ", ".join(sorted(dup.unique())[:20])
        )

    direct_map = {
        acc: idx
        for idx, acc in meta["archive_data_accession_norm"].items()
        if pd.notna(acc) and acc != ""
    }

    # Build accession -> metadata-row map from FASTQ download paths.
    link_map: dict[str, list[int]] = {}
    for idx, links in meta["download_links"].items():
        for acc in accessions_from_download_links(links):
            link_map.setdefault(acc, []).append(idx)

    rows = []
    report_rows = []

    for _, r in results.iterrows():
        analysis_acc = r["analysis_accession_norm"]
        match_idx = None
        method = None

        if analysis_acc in direct_map:
            match_idx = direct_map[analysis_acc]
            method = "direct_accession"
        else:
            candidates = sorted(set(link_map.get(analysis_acc, [])))
            if len(candidates) == 1:
                match_idx = candidates[0]
                method = "download_link_accession"
            elif len(candidates) > 1:
                raise ValueError(
                    f"Ambiguous download-link match for {analysis_acc}: "
                    f"metadata rows {candidates}"
                )

        if match_idx is None:
            report_rows.append({
                "analysis_accession": analysis_acc,
                "status": "unmatched",
                "match_method": "",
                "AncientMetagenomeDir_accession": "",
                "project_name": "",
            })
            continue

        m = meta.loc[match_idx]
        amd_acc = str(m["archive_data_accession"]).strip()

        note = ""
        if method == "download_link_accession" and analysis_acc != amd_acc.upper():
            note = (
                "Analysis accession derived from FASTQ download filename; "
                "differs from archive_data_accession recorded in AncientMetagenomeDir."
            )

        rows.append({
            "analysis_accession": str(r["sample"]).strip(),
            "AncientMetagenomeDir_accession": amd_acc,
            "match_method": method,
            "project_name": m["project_name"],
            "publication_year": m["publication_year"],
            "publication_doi": m["data_publication_doi"],
            "sample_name": m["sample_name"],
            "archive": m["archive"],
            "archive_project": m["archive_project"],
            "archive_sample_accession": m["archive_sample_accession"],
            "library_name": m["library_name"],
            "strand_type": m["strand_type"],
            "library_polymerase": m["library_polymerase"],
            "library_treatment": m["library_treatment"],
            "library_concentration": m["library_concentration"],
            "instrument_model": m["instrument_model"],
            "library_layout": m["library_layout"],
            "library_strategy": m["library_strategy"],
            "read_used": read_used(m["library_layout"]),
            "reported_read_count": m["read_count"],
            "best_length_bp": r["best_length"],
            "max_mean_period3_R2": r["max_r2"],
            "download_links": m["download_links"],
            "download_md5s": m["download_md5s"],
            "download_sizes": m["download_sizes"],
            "notes": note,
        })

        report_rows.append({
            "analysis_accession": analysis_acc,
            "status": "matched",
            "match_method": method,
            "AncientMetagenomeDir_accession": amd_acc,
            "project_name": m["project_name"],
        })

    table = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)
    report = pd.DataFrame(report_rows)

    # Stable, publication-friendly order.
    table = table.sort_values(
        ["project_name", "analysis_accession"],
        kind="stable",
        na_position="last",
    ).reset_index(drop=True)

    return table, report


def write_excel(
    table: pd.DataFrame,
    xlsx_path: Path,
    metadata_path: Path,
    results_path: Path,
    report: pd.DataFrame,
) -> None:
    readme_rows = [
        ["Supplementary Table S1",
         "Dataset provenance and selected period-3 summary values for the ancient-DNA cross-library analysis."],
        ["Table size", f"{len(table):,} analyzed libraries"],
        ["Direct accession matches",
         str((table["match_method"] == "direct_accession").sum())],
        ["Download-link-derived matches",
         str((table["match_method"] == "download_link_accession").sum())],
        ["Unmatched analysis rows",
         str((report["status"] == "unmatched").sum())],
        ["AncientMetagenomeDir source file", metadata_path.name],
        ["AncientMetagenomeDir source SHA-256", sha256(metadata_path)],
        ["Period-3 results source file", results_path.name],
        ["Period-3 results source SHA-256", sha256(results_path)],
        ["Matching note",
         "Direct matches use archive_data_accession. If absent, the analysis accession is matched to an accession embedded in download_links. Both identifiers are retained when they differ."],
        ["Analysis note",
         "For paired-end datasets, R1 was used for read-coordinate composition analysis; single-end datasets were analyzed directly."],
        ["Figure statistic",
         "For each library, the eligible exact read length with the largest mean period-3 R² across A/T/G/C is reported as max_mean_period3_R2."],
        ["AncientMetagenomeDir version",
         "Historical Git release/commit was not recoverable; the exact metadata TSV used here is identified by filename and SHA-256 and should be archived with the manuscript code/data release."],
        ["", ""],
        ["Column", "Description"],
    ]
    readme_rows.extend([[c, COLUMN_DESCRIPTIONS[c]] for c in OUTPUT_COLUMNS])
    readme = pd.DataFrame(readme_rows, columns=["Item", "Description"])

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        readme.to_excel(writer, sheet_name="README", index=False)
        table.to_excel(writer, sheet_name="Supplementary Table S1", index=False)

    wb = load_workbook(xlsx_path)

    # README formatting
    ws = wb["README"]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 110
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    # Highlight the column dictionary header.
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == "Column" and ws.cell(row, 2).value == "Description":
            for cell in ws[row]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")

    # Main table formatting
    ws = wb["Supplementary Table S1"]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "analysis_accession": 20,
        "AncientMetagenomeDir_accession": 28,
        "match_method": 24,
        "project_name": 20,
        "publication_year": 14,
        "publication_doi": 28,
        "sample_name": 22,
        "archive": 10,
        "archive_project": 18,
        "archive_sample_accession": 24,
        "library_name": 30,
        "strand_type": 16,
        "library_polymerase": 24,
        "library_treatment": 22,
        "library_concentration": 22,
        "instrument_model": 26,
        "library_layout": 16,
        "library_strategy": 18,
        "read_used": 16,
        "reported_read_count": 20,
        "best_length_bp": 16,
        "max_mean_period3_R2": 22,
        "download_links": 70,
        "download_md5s": 70,
        "download_sizes": 28,
        "notes": 60,
    }
    header_to_col = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for name, width in widths.items():
        ws.column_dimensions[get_column_letter(header_to_col[name])].width = width

    # Formats and wrapping.
    r2_col = header_to_col["max_mean_period3_R2"]
    count_col = header_to_col["reported_read_count"]
    length_col = header_to_col["best_length_bp"]
    pubyear_col = header_to_col["publication_year"]

    for row in range(2, ws.max_row + 1):
        ws.cell(row, r2_col).number_format = "0.0000"
        ws.cell(row, count_col).number_format = "#,##0"
        ws.cell(row, length_col).number_format = "0"
        ws.cell(row, pubyear_col).number_format = "0"
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)

        # Visually flag the three accession discrepancies for review.
        if ws.cell(row, header_to_col["match_method"]).value == "download_link_accession":
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = PatternFill("solid", fgColor="FFF2CC")

    wb.save(xlsx_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--metadata",
        type=Path,
        required=True,
        help="AncientMetagenomeDir environmental-library TSV",
    )
    parser.add_argument(
        "--results",
        type=Path,
        required=True,
        help="best_length_per_library.csv",
    )
    parser.add_argument(
        "--outdir",
        type=Path,
        default=Path("supplementary"),
    )
    parser.add_argument(
        "--prefix",
        default="Supplementary_Table_S1",
    )
    args = parser.parse_args()

    args.outdir.mkdir(parents=True, exist_ok=True)

    meta = pd.read_csv(args.metadata, sep="\t", low_memory=False)
    results = pd.read_csv(args.results)

    table, report = build_table(meta, results)

    tsv_path = args.outdir / f"{args.prefix}.tsv"
    xlsx_path = args.outdir / f"{args.prefix}.xlsx"
    report_path = args.outdir / f"{args.prefix}_match_report.tsv"

    table.to_csv(tsv_path, sep="\t", index=False, na_rep="")
    report.to_csv(report_path, sep="\t", index=False, na_rep="")
    write_excel(table, xlsx_path, args.metadata, args.results, report)

    n_direct = int((table["match_method"] == "direct_accession").sum())
    n_link = int((table["match_method"] == "download_link_accession").sum())
    n_unmatched = int((report["status"] == "unmatched").sum())

    print(f"Wrote: {xlsx_path}")
    print(f"Wrote: {tsv_path}")
    print(f"Wrote: {report_path}")
    print(f"Libraries in results: {len(results):,}")
    print(f"Matched table rows: {len(table):,}")
    print(f"  direct accession matches: {n_direct:,}")
    print(f"  download-link-derived matches: {n_link:,}")
    print(f"  unmatched: {n_unmatched:,}")

    if len(table) != len(results) or n_unmatched:
        raise SystemExit(
            "ERROR: not every analysis row was matched. Inspect the match report."
        )


if __name__ == "__main__":
    main()
